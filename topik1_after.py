import pandas as pd
from openpyxl import Workbook
wb = Workbook()
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_excel('d:/python/study/TOPIK/data/topik1_after.xlsx', encoding='euckr')
a = pd.value_counts(df['태소'].values, sort=True)
#for v in a.index :
#    전체 데이터 프레임을 for문으로 돌려서 키 값과 일치하면 새로운 프레임에 태소를 추가
inf_dict = {}
#for i in range(len(df)) :
#    inf = df.iloc[i].원형
#    pum = df.iloc[i].품사
#    tae = df.iloc[i].태소
#    mun = df.iloc[i].문장
#    if inf not in inf_dict.keys() :
#        inf_dict[inf] = [pum, tae]
#    if inf in inf_dict.keys() :
#        if tae in inf_dict[inf] :
#            continue
#        elif len(inf_dict[inf]) < 6 :
#            inf_dict[inf].append(tae)
#        elif len(inf_dict[inf]) > 5 :
#            for l in range(1,5) :
#                if a[tae] > a[inf_dict[inf][1:]].values.any() :
#                    if a[inf_dict[inf][l]] < a[tae] :
#                        inf_dict[inf][l] = tae
#                        break
for i in range(len(df)) :
    inf = df.iloc[i].원형
    pum = df.iloc[i].품사
    tae = df.iloc[i].태소
    mun = df.iloc[i].문장
    if inf not in inf_dict.keys() :
        inf_dict[inf] = [pum, mun]
    if inf in inf_dict.keys() :
        if mun in inf_dict[inf] :
            continue
        elif len(inf_dict[inf]) < 6 :
            inf_dict[inf].append(mun)
ws1 = wb.create_sheet(title = "inf_sen")
ws2 = wb.create_sheet(title = "inf_li")
k_list = []
#####
for k, v in inf_dict.items() :
    ws1.append(v)
key_list = []
for kl in inf_dict.keys() :
    key_list.append(kl)

k_list_df = pd.DataFrame(key_list)
for row in dataframe_to_rows(k_list_df, index=True, header=True):
    if len(row) > 1:
        ws2.append(row)
#df_after = pd.DataFrame(inf_dict, index = pd.series["1", "2"], columns = ["1", "2", "3", "4", "5"])
#print(df_after)
wb.save("D:/python/study/topik/excel_test.xlsx") #저장
#        inf_series = pd.Series(tae, index=[inf_dict])
#print(inf_dict['가게'])
#print(inf_series)
#for a in a :
#    print(a)
#q = list(range(5))
#w = list(range(5))
#s = pd.Series(q, index=[w])
#print(s)
