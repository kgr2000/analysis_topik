#import sys
#import io
#sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
#sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')
import codecs
import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
wb = Workbook()
import matplotlib.pyplot as plt
from konlpy.tag import Okt
from konlpy.utils import pprint

okt = Okt()
word_dic_t = {}
word_dic_n = {}
word_dic_v = {}
word_dic_a = {}
wd_ct_t_et = 0
wd_ct = 0
wd_ct_t = 0
wd_ct_n = 0
wd_ct_v = 0
wd_ct_a = 0
# 끝

lines = list()
txt = codecs.open('topik1_test.txt', 'r', encoding='utf-8')
for line in txt :
    fletter = re.sub("(\(\S+\))", "", line)
    fletter = re.sub("(\(\S+)", "", fletter)
    fletter = re.sub("(\S+\))", "", fletter)
    fletter = re.sub("제\S+회", "", fletter)
    fletter = re.sub("한국\S+시험", "", fletter)
    fletter = re.sub("\S형", "", fletter)
    fletter = re.sub("남자 :", "", fletter)
    fletter = re.sub("여자 :", "", fletter)
#    if "시오." in line :
#        continue
    if "※" in line :
        continue
    elif "능력시험" in line :
        continue
    elif "일반" in line :
        continue
    elif "<" in line :
        continue
    elif "TOPIK" in line :
        continue
    elif "～" in line :
        continue
    elif "B" in line :
        continue
#    elif "나 " in line :
#        print(line)
    lines.append(fletter)
#for line in lines :
#    print(line)
ws_t = wb.create_sheet(title = "T1_Tbw")
ws_n = wb.create_sheet(title = "T1_Nbw")
ws_v = wb.create_sheet(title = "T1_Vbw")
ws_a = wb.create_sheet(title = "T1_Abw")
for line in lines :
    #품사 분류
    wd = okt.pos(line) #, norm=True, stem=True
    for taeso, pumsa in wd :
        wd_ct_t_et += 1
        if pumsa != "Punctuation" and pumsa != "Foreign" and pumsa != "Number" and pumsa != "Unknown" and pumsa != "Alpha" and pumsa != "Josa" :
            t = okt.pos(taeso, norm=True, stem=True)
            for aa, bb in t :
                t_list = [aa, taeso, pumsa]
                ws_t.append(t_list)
            for taeso, pumsa in t :
                if not (taeso in word_dic_t) :
                    word_dic_t[taeso] = 0
                word_dic_t[taeso] += 1
                wd_ct_t += 1
        if pumsa == "Noun" :
            n = okt.pos(taeso, norm=True, stem=True)
            for aa, bb in n :
                n_list = [aa, taeso, pumsa]
                ws_n.append(n_list)
            for taeso, pumsa in n :
                if not (taeso in word_dic_n) :
                    word_dic_n[taeso] = 0
                word_dic_n[taeso] += 1
                wd_ct_n += 1
        if pumsa == "Verb" :
            v = okt.pos(taeso, norm=True, stem=True)
            for aa, bb in v :
                v_list = [aa, taeso, pumsa]
                ws_v.append(v_list)
            for taeso, pumsa in v :
                if not (taeso in word_dic_v) :
                    word_dic_v[taeso] = 0
                word_dic_v[taeso] += 1
                wd_ct_v += 1
        if pumsa == "Adjective" :
            a = okt.pos(taeso, norm=True, stem=True)
            for aa, bb in a :
                a_list = [aa, taeso, pumsa]
                ws_a.append(a_list)
            for taeso, pumsa in a :
                if not (taeso in word_dic_a) :
                    word_dic_a[taeso] = 0
                word_dic_a[taeso] += 1
                wd_ct_a += 1
#print(a_list)
#for wd_a in word_dic_a.keys() :
#    wd_a = okt.pos(wd_a, norm=True, stem=True)
#    print(wd_a)
#print(okt.pos(wd_a))
#print(okt.pos(word_dic_a.keys(), norm=True, stem=True))
#for wd_a, v in word_dic_a :
#    wd_a = okt.pos(wd_a, norm=True, stem=True)
#    print(wd_a)
    ##해당 키워드 검색###
#            if taeso == "홀" :
#                print(line)
    ##해당 키워드 검색 끝###
    #빈도 순서에 따라 정렬#
keys_t = sorted(word_dic_t.items(), key=lambda x:x[1], reverse=True)
keys_n = sorted(word_dic_n.items(), key=lambda x:x[1], reverse=True)
keys_v = sorted(word_dic_v.items(), key=lambda x:x[1], reverse=True)
keys_a = sorted(word_dic_a.items(), key=lambda x:x[1], reverse=True)
#엑셀 Sheet 생성/자료 입력
ws_tt = wb.create_sheet(title = "T1_Tbc")
for row_t in keys_t :
    ws_tt.append(row_t)
ws_nn = wb.create_sheet(title = "T1_Nbc")
for row_n in keys_n :
    ws_nn.append(row_n)
ws_vv = wb.create_sheet(title = "T1_Vbc")
for row_v in keys_v :
    ws_vv.append(row_v)
ws_aa = wb.create_sheet(title = "T1_Abc")
for row_a in keys_a :
    ws_aa.append(row_a)
wb.save("D:/python/study/excel_test.xlsx") #저장
#count_t = 0
#count_w = 0
#d_wd = []
#d_ct = []
#keys_t_input = 10
#for word, count in keys_t[:keys_t_input] :
#    count_t += count
#total_count = count_t
#count_t = 0

#for word, count in keys_t[:keys_t_input] :
#    count_w += 1
#    count_t += count
#    d_wd.append(count_w)
#    d_ct.append(count_t/total_count*100)
#data = {"number": d_wd,
#        "covers": d_ct}
#df = pd.DataFrame(data)
#print(df)
#plt.plot(d_wd, d_ct)
#plt.show()
#    print(word, count)

#print(keys_t[:keys_t_input])
#print("\n최다 사용 단어 상위", keys_t_input, "개 전체 커버율: ", count_t/wd_ct*100, "%")
    #print("{0}({1})".format(word, count), end=" ")
#for word, count in keys_v[:10] :
#    print("{0}({1})".format(word, count), end=" ")
#for word, count in keys_a[:10] :
#    print("{0}({1})".format(word, count), end=" ")

#리스트의 길이, 튜플의 길이 총 사용된 단어 수

#print("역대 TOPIK 초급 단어 분석")
#print("전체 문자열 수: ", wd_ct_t_et, "개")
#print("총 단어 사용 횟수: ", wd_ct, "회")
#print("총 단어 수: ", len(keys_t), "개")
#print("총 명사 단어 수: ", len(keys_n), "개")
#print("총 동사 단어 수: ", len(keys_v), "개")
#print("총 형용사 단어 수: ", len(keys_a), "개")
#print("총 명사 사용 횟수: ", wd_ct_n, "회")
#print("총 동사 사용 횟수: ", wd_ct_v, "회")
#print("총 형용사 사용 횟수: ", wd_ct_a, "회")
    #끝
# 엑셀 관련
